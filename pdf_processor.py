import re
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List, Dict, Optional
import logging
import PyPDF2
import os
import pandas as pd
from io import BytesIO
from image_processor import ImageProcessor

logger = logging.getLogger(__name__)

class PDFProcessor:
    def __init__(self):
        self.processed_data = []
        self.image_processor = ImageProcessor()
    
    def process_pdf(self, file_path: str) -> Dict:
        """
        Основной метод обработки PDF файлов
        Всегда использует Claude API для максимальной точности извлечения данных
        """
        try:
            logger.info(f"Начинаю обработку PDF с Claude API: {file_path}")
            
            # Проверяем доступность Claude API
            if hasattr(self.image_processor, 'claude_client') and self.image_processor.claude_client:
                logger.info("Использую Claude API для обработки PDF")
                return self._process_scanned_pdf(file_path)
            else:
                logger.warning("Claude API недоступен, использую стандартный метод")
                # Fallback на стандартный метод только если Claude недоступен
                text_result = self._extract_text_standard(file_path)
                
                # Если и стандартный метод не дал результата
                if not text_result['text'] or len(text_result['text'].strip()) < 10:
                    return {
                        'success': False,
                        'text': '',
                        'pages': 0,
                        'tables': [],
                        'method': 'Error',
                        'error': 'Claude API недоступен и стандартный метод не извлек данные'
                    }
                
                return text_result
            
        except Exception as e:
            logger.error(f"Ошибка при обработке PDF {file_path}: {e}")
            return {
                'success': False,
                'text': '',
                'pages': 0,
                'tables': [],
                'method': 'Error',
                'error': str(e)
            }

    def _extract_text_standard(self, file_path: str) -> Dict:
        """
        Стандартное извлечение текста из PDF
        """
        try:
            text = ""
            pages = 0
            
            # Пробуем с PyPDF2
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                pages = len(pdf_reader.pages)
                
                for page_num, page in enumerate(pdf_reader.pages):
                    page_text = page.extract_text()
                    if page_text:
                        text += f"\n\n--- Страница {page_num + 1} ---\n{page_text}"
            
            # Если PyPDF2 не дал результата, пробуем pdfplumber
            if not text.strip():
                logger.info("PyPDF2 не извлек текст, пробую pdfplumber")
                with pdfplumber.open(file_path) as pdf:
                    pages = len(pdf.pages)
                    for page_num, page in enumerate(pdf.pages):
                        page_text = page.extract_text()
                        if page_text:
                            text += f"\n\n--- Страница {page_num + 1} ---\n{page_text}"
            
            # Извлекаем таблицы
            tables = self.extract_table_from_pdf(file_path)
            
            return {
                'success': True,
                'text': text.strip(),
                'pages': pages,
                'tables': tables,
                'method': 'Standard PDF extraction',
                'error': None
            }
            
        except Exception as e:
            logger.error(f"Ошибка стандартного извлечения из {file_path}: {e}")
            return {
                'success': False,
                'text': '',
                'pages': 0,
                'tables': [],
                'method': 'Standard PDF extraction',
                'error': str(e)
            }

    def _process_scanned_pdf(self, file_path: str) -> Dict:
        """
        Обработка PDF с помощью Claude API OCR
        """
        try:
            logger.info(f"Обрабатываю PDF с Claude API: {file_path}")
            result = self.image_processor.process_pdf_images(file_path)
            
            # Добавляем информацию о методе обработки
            if result['success']:
                result['method'] = f"Claude API OCR ({result.get('method', 'Unknown')})"
                tables_count = len(result.get('tables', []))
                text_length = len(result.get('text', ''))
                logger.info(f"PDF обработан Claude API: метод={result['method']}, таблиц={tables_count}, символов={text_length}")
            else:
                logger.error(f"Ошибка Claude API обработки: {result.get('error', 'Unknown')}")
            
            return result
            
        except Exception as e:
            logger.error(f"Ошибка обработки сканированного PDF {file_path}: {e}")
            return {
                'success': False,
                'text': '',
                'pages': 0,
                'tables': [],
                'method': 'OCR Error',
                'error': str(e)
            }

    def _extract_text_from_pdf(self, pdf_path: str) -> str:
        """
        Извлекает весь текст из PDF файла
        """
        text = ""
        
        try:
            # Пробуем с pdfplumber (лучше для таблиц)
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n\n"
        except Exception as e:
            logger.warning(f"pdfplumber не смог обработать файл: {e}")
            
            # Если pdfplumber не работает, пробуем PyPDF2
            try:
                with open(pdf_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    for page in pdf_reader.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text += page_text + "\n\n"
            except Exception as e2:
                logger.error(f"PyPDF2 также не смог обработать файл: {e2}")
                raise e2
        
        return text.strip()
    
    def _count_pages(self, pdf_path: str) -> int:
        """
        Подсчитывает количество страниц в PDF
        """
        try:
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                return len(pdf_reader.pages)
        except Exception as e:
            logger.error(f"Ошибка при подсчете страниц: {e}")
            return 0

    def extract_table_from_pdf(self, file_path: str) -> List[Dict]:
        """
        Извлекает таблицы из PDF файла
        """
        tables = []
        try:
            with pdfplumber.open(file_path) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    page_tables = page.extract_tables()
                    for table in page_tables:
                        if table and len(table) > 1:  # Есть данные и заголовки
                            headers = table[0]
                            for row in table[1:]:
                                if len(row) >= len(headers):
                                    table_dict = {}
                                    for i, header in enumerate(headers):
                                        if header and i < len(row):
                                            # Определяем тип данных по заголовку
                                            header_lower = str(header).lower()
                                            if 'фио' in header_lower or 'имя' in header_lower:
                                                table_dict['fio'] = row[i]
                                            elif 'дата' in header_lower and 'рожд' in header_lower:
                                                table_dict['birth_date'] = row[i]
                                            elif 'должн' in header_lower or 'профес' in header_lower:
                                                table_dict['position'] = row[i]
                                            elif 'вред' in header_lower or 'фактор' in header_lower:
                                                table_dict['harmful_factors'] = row[i]
                                            else:
                                                if 'other' not in table_dict:
                                                    table_dict['other'] = ''
                                                table_dict['other'] += f"{header}: {row[i]}; "
                                    
                                    if table_dict:  # Если есть данные
                                        tables.append(table_dict)
        
        except Exception as e:
            logger.error(f"Ошибка извлечения таблиц из {file_path}: {e}")
        
        return tables
    
    def _find_headers(self, table: List[List]) -> Optional[int]:
        """
        Находит строку с заголовками таблицы
        """
        for i, row in enumerate(table):
            if row and any(cell for cell in row if cell and ('ФИО' in str(cell) or 'фио' in str(cell).lower())):
                return i
        return None
    
    def _parse_table_data(self, table: List[List], header_row_idx: int) -> List[Dict]:
        """
        Парсит данные из таблицы
        """
        data = []
        headers = table[header_row_idx] if header_row_idx is not None else None
        
        if not headers:
            return data
        
        # Нормализуем заголовки
        normalized_headers = []
        for header in headers:
            if header:
                header_str = str(header).strip()
                if 'ФИО' in header_str or 'фио' in header_str.lower():
                    normalized_headers.append('fio')
                elif 'дата' in header_str.lower() and ('рож' in header_str.lower() or 'birth' in header_str.lower()):
                    normalized_headers.append('birth_date')
                elif 'должность' in header_str.lower() or 'position' in header_str.lower():
                    normalized_headers.append('position')
                elif 'фактор' in header_str.lower() or 'вредн' in header_str.lower():
                    normalized_headers.append('harmful_factors')
                else:
                    normalized_headers.append('other')
            else:
                normalized_headers.append('empty')
        
        # Парсим строки данных
        for row in table[header_row_idx + 1:]:
            if row and any(cell for cell in row if cell and str(cell).strip()):
                row_data = {}
                for i, cell in enumerate(row):
                    if i < len(normalized_headers):
                        header = normalized_headers[i]
                        if header != 'empty' and cell:
                            row_data[header] = str(cell).strip()
                
                if row_data:
                    data.append(row_data)
        
        return data
    
    def _extract_from_text(self, text: str) -> List[Dict]:
        """
        Извлекает данные из текста, если таблица не найдена
        """
        data = []
        lines = text.split('\n')
        
        # Паттерны для поиска данных
        fio_pattern = r'([А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+)'
        date_pattern = r'(\d{2}\.\d{2}\.\d{4})'
        
        current_entry = {}
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # Поиск ФИО
            fio_matches = re.findall(fio_pattern, line)
            if fio_matches:
                if current_entry:
                    data.append(current_entry)
                current_entry = {'fio': fio_matches[0]}
            
            # Поиск даты рождения
            date_matches = re.findall(date_pattern, line)
            if date_matches and current_entry:
                current_entry['birth_date'] = date_matches[0]
            
            # Поиск должности
            if any(word in line.lower() for word in ['рабочий', 'инженер', 'врач', 'медсестра']):
                if current_entry:
                    current_entry['position'] = line
        
        if current_entry:
            data.append(current_entry)
        
        return data
    
    def create_excel_from_data(self, data: List[Dict], output_path: str) -> str:
        """
        Создает Excel файл из извлеченных данных
        """
        if not data:
            raise ValueError("Нет данных для создания Excel файла")
        
        # Создаем новую рабочую книгу
        wb = Workbook()
        ws = wb.active
        
        # Определяем тип данных и название листа
        is_organization = any('organization_name' in row or 'inn' in row for row in data)
        ws.title = "Организации" if is_organization else "Медицинские осмотры"
        
        # Определяем колонки в зависимости от типа документа
        if is_organization:
            column_mapping = {
                'organization_name': 'Название организации',
                'inn': 'ИНН',
                'kpp': 'КПП',
                'address': 'Адрес',
                'director': 'Директор',
                'phone': 'Телефон',
                'email': 'Email',
                'license': 'Лицензия',
                'other': 'Дополнительная информация'
            }
        else:
            column_mapping = {
                'fio': 'ФИО',
                'birth_date': 'Дата рождения',
                'position': 'Должность',
                'department': 'Отдел',
                'harmful_factors': 'Факторы вредности',
                'other': 'Дополнительная информация'
            }
        
        # Находим все уникальные ключи в данных
        all_keys = set()
        for row in data:
            all_keys.update(row.keys())
        
        # Создаем заголовки
        headers = []
        if is_organization:
            priority_keys = ['organization_name', 'inn', 'kpp', 'address', 'director', 'phone', 'email', 'license']
        else:
            priority_keys = ['fio', 'birth_date', 'position', 'department', 'harmful_factors']
        
        for key in priority_keys:
            if key in all_keys:
                headers.append(column_mapping.get(key, key))
        
        # Добавляем остальные ключи
        for key in all_keys:
            if key not in priority_keys:
                headers.append(column_mapping.get(key, key.title()))
        
        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Записываем данные
        other_keys = [k for k in all_keys if k not in priority_keys]
        
        for row_idx, row_data in enumerate(data, 2):
            col_idx = 1
            
            # Записываем основные колонки в порядке
            for key in priority_keys:
                if key in all_keys:
                    value = row_data.get(key, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    col_idx += 1
            
            # Записываем дополнительные колонки
            for key in other_keys:
                value = row_data.get(key, "")
                ws.cell(row=row_idx, column=col_idx, value=value)
                col_idx += 1
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл
        wb.save(output_path)
        return output_path

    def create_excel_from_tables(self, tables: List[Dict], output_path: str) -> bool:
        """
        Создает Excel файл из извлеченных таблиц
        """
        try:
            if not tables:
                logger.warning("Нет таблиц для экспорта в Excel")
                return False
            
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Извлеченные данные"
            
            # Заголовки
            headers = ["ФИО", "Дата рождения", "Должность", "Факторы вредности", "Прочее"]
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)
            
            # Данные
            row = 2
            for table_data in tables:
                worksheet.cell(row=row, column=1, value=table_data.get('fio', ''))
                worksheet.cell(row=row, column=2, value=table_data.get('birth_date', ''))
                worksheet.cell(row=row, column=3, value=table_data.get('position', ''))
                worksheet.cell(row=row, column=4, value=table_data.get('harmful_factors', ''))
                worksheet.cell(row=row, column=5, value=table_data.get('other', ''))
                row += 1
            
            workbook.save(output_path)
            logger.info(f"Excel файл создан: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка создания Excel файла: {e}")
            return False 